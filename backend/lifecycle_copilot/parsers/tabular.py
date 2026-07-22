import csv
import io
import re
from typing import Any, Optional


SUPPORTED_FORMATS = {"csv", "xlsx"}


def detect_tabular_format(filename: str) -> Optional[str]:
    lowered = (filename or "").strip().lower()
    if lowered.endswith(".csv"):
        return "csv"
    if lowered.endswith(".xlsx"):
        return "xlsx"
    return None


def parse_tabular_upload(filename: str, raw: bytes) -> tuple[list[str], list[list[str]]]:
    file_format = detect_tabular_format(filename)
    if file_format == "csv":
        return _parse_csv(raw)
    if file_format == "xlsx":
        return _parse_xlsx(raw)
    raise ValueError("unsupported_format")


def rows_to_dicts(headers: list[str], rows: list[list[str]]) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    for row in rows:
        item: dict[str, str] = {}
        for index, header in enumerate(headers):
            item[header] = row[index].strip() if index < len(row) and row[index] is not None else ""
        records.append(item)
    return records


def _parse_csv(raw: bytes) -> tuple[list[str], list[list[str]]]:
    text = raw.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [list(row) for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        raise ValueError("empty_file")
    headers = [header.strip() for header in rows[0]]
    return headers, rows[1:]


def _parse_xlsx(raw: bytes) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("xlsx_support_unavailable") from exc

    workbook = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        values = [
            "" if value is None else str(value).strip()
            for value in row
        ]
        if any(values):
            rows.append(values)
    workbook.close()

    if not rows:
        raise ValueError("empty_file")

    headers = rows[0]
    return headers, rows[1:]


def normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (value or "").strip().lower()).strip("_")
